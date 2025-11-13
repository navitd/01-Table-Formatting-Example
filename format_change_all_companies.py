# in the csv file - data file - take everything. column z is the one that decides which rows are combined
# groupby column z
# in the second file, file "DCF v6" I need to populate up to column T






import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import os

def process_company_data(input_csv, company_name, licensee_name, province, sheet_name, workbook_name):
    """Process one company CSV file and add a formatted worksheet to the given Excel workbook."""
    
    # Define column mappings and settings
    use_cols = ["Tier-4 id", "record_id", "area_name*", "latitude", "longitude",
                "structure_height", "tx_ant_azimuth", "licence_category*",
                "technology", "tx_power", "bandwidth"]
    group_unique_values = ['Tier-4 id', 'record_id', 'latitude', 'longitude', 'structure_height']
    group_repeated_values = ['licence_category*', 'tx_ant_azimuth', 'technology', 'bandwidth', 'tx_power']
    name_groupby = 'record_id'

    # Read input CSV
    df = pd.read_csv(input_csv, usecols=use_cols)

    # Group and aggregate
    results = []
    for group_name, group_df in df.groupby(name_groupby):
        row_result = {name_groupby: group_name}
        for col in group_unique_values:
            row_result[col] = group_df.iloc[0][col]
        for col in group_repeated_values:
            row_result[col] = '~'.join(group_df[col].astype(str).tolist())

        # Initialize accumulators
        row_result['4G_BW'] = 0
        row_result['5G_BW'] = 0
        for _, row in group_df.iterrows():
            tech = str(row['technology'])
            bw = float(row['bandwidth'])
            if tech == 'LTE':
                row_result['4G_BW'] += bw
            elif tech == '5G':
                row_result['5G_BW'] += bw 

        results.append(row_result)

    results = pd.DataFrame(results) #to be printed in the excel

    # Rename columns to destination names
    results = results.rename(columns={
        'Tier-4 id': 'Code',
        'record_id': 'Tower ID',
        'latitude': 'LATITUDE',
        'longitude': 'LONGITUDE',
        'structure_height': 'Height (m)',
        'tx_ant_azimuth': 'TX_ANT_AZI',
        'licence_category*': 'SERVICES',
        'technology': 'TECHNOLOGY',
        'tx_power': 'TX_PWR',
        'bandwidth': 'TR_BW_BLOCS'
    })

    # Add and transform fields
    results["Name"] = ' '
    results["Province"] = province
    results["Licensee"] = licensee_name
    results["Type"] = results["Height (m)"].apply(lambda x: 'Micro' if x < 10 else 'Macro')
    results["Micro"] = results["Type"].apply(lambda x: 1 if x == 'Micro' else '-')
    results["Macro"] = results["Type"].apply(lambda x: 1 if x == 'Macro' else '-')
    results["SERVICES"] = results["SERVICES"].str.replace("~", "|")
    results["TECHNOLOGY"] = results["TECHNOLOGY"].str.replace("LTE", "4G")

    # Reorder columns
    final_columns = ["Code", "Name", "Province", "Tower ID", "Licensee", "LATITUDE", "LONGITUDE",
                     "Height (m)", "Type", "Micro", "Macro", "TX_ANT_AZI", "SERVICES",
                     "TECHNOLOGY", "TR_BW_BLOCS", "4G_BW", "5G_BW", "TX_PWR"]
    results = results[final_columns].copy()

    # Append or create workbook
    if os.path.exists(workbook_name):
        with pd.ExcelWriter(workbook_name, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            results.to_excel(writer, index=False, sheet_name=sheet_name)
    else:
        with pd.ExcelWriter(workbook_name, engine="openpyxl") as writer:
            results.to_excel(writer, index=False, sheet_name=sheet_name)

    # Apply styling
    wb = load_workbook(workbook_name)
    ws = wb[sheet_name]
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # dark blue
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    wb.save(workbook_name)

    print(f"✅ Added worksheet '{sheet_name}' for {company_name} to {workbook_name}")


################################################################            main           ################################################################

output_workbook = "Telecom_Data_format2.xlsx"

process_company_data(
    input_csv="Rogers sectors v3 Final.csv",
    company_name="Rogers",
    licensee_name="Rogers Communications Canada Inc.",
    province="QC",
    sheet_name="Rogers",
    workbook_name=output_workbook
)

process_company_data(
    input_csv="TELUS sectors v3 Final.csv",
    company_name="TELUS",
    licensee_name="TELUS",
    province="QC",
    sheet_name="TELUS",
    workbook_name=output_workbook
)


process_company_data(
    input_csv="Videotron sectors v3 Final.csv",
    company_name="Videotron",
    licensee_name="Videotron.",
    province="QC",
    sheet_name="Videotron",
    workbook_name=output_workbook
)







